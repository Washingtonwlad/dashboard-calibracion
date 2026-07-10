from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

C_DARK = "22194E"
C_ACCENT = "FF4298"
C_ORANGE = "FFAB48"
C_LIGHT = "F4F0FA"
C_GREEN = "4A9E6B"
C_RED = "C95F5F"
C_WHITE = "FFFFFF"


def _clasificacion(valor):
    if valor >= 85:
        return "Adecuado"
    if valor >= 70:
        return "Cercano"
    return "Alejado"


def _estilizar_tabla(ws, fila_encabezado, fila_final, columna_final):
    fill = PatternFill("solid", fgColor=C_DARK)
    borde = Border(bottom=Side(style="thin", color="D9D1E8"))
    for cell in ws[fila_encabezado]:
        if cell.column > columna_final:
            break
        cell.fill = fill
        cell.font = Font(color=C_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(
        min_row=fila_encabezado + 1,
        max_row=fila_final,
        min_col=1,
        max_col=columna_final,
    ):
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(vertical="center")
    ws.auto_filter.ref = f"A{fila_encabezado}:{get_column_letter(columna_final)}{fila_final}"
    ws.freeze_panes = f"A{fila_encabezado + 1}"
    ws.sheet_view.showGridLines = False


def _ajustar_anchos(ws, max_width=38):
    for columna in range(1, ws.max_column + 1):
        letra = get_column_letter(columna)
        longitud = max(
            (len(str(ws.cell(fila, columna).value or "")) for fila in range(1, ws.max_row + 1)),
            default=0,
        )
        ws.column_dimensions[letra].width = min(max(longitud + 2, 11), max_width)


def _escribir_titulo(ws, titulo, subtitulo=None):
    ws["A1"] = titulo
    ws["A1"].font = Font(size=18, bold=True, color=C_DARK)
    if subtitulo:
        ws["A2"] = subtitulo
        ws["A2"].font = Font(size=10, color="777777")


def crear_excel_resultados(
    meta,
    df_base,
    df_actual,
    df_simulado,
    competencias,
    esperados_actuales,
    esperados_simulados,
    pesos_simulados,
):
    """Genera el entregable Excel de la calibración actual y simulada."""
    wb = Workbook()

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    _escribir_titulo(ws, "Calibración de perfil", meta.get("Nombre del Proceso", ""))
    fila = 4
    for etiqueta, clave in [
        ("Perfil", "Nombre del Perfil"),
        ("Proceso", "Nombre del Proceso"),
        ("Inicio", "Inicio"),
        ("Fin", "Fin"),
        ("Reclutador", "Reclutador"),
    ]:
        if meta.get(clave):
            ws.cell(fila, 1, etiqueta).font = Font(bold=True, color=C_DARK)
            ws.cell(fila, 2, meta[clave])
            fila += 1

    fila += 1
    ws.cell(fila, 1, "Indicador")
    ws.cell(fila, 2, "Actual")
    ws.cell(fila, 3, "Simulado")
    ws.cell(fila, 4, "Variación")
    encabezado = fila
    metricas = [
        ("CAP promedio", df_actual["CAP_global"].mean(), df_simulado["CAP_global"].mean()),
        ("Candidatos adecuados", (df_actual["CAP_global"] >= 85).sum(), (df_simulado["CAP_global"] >= 85).sum()),
        ("Candidatos cercanos", ((df_actual["CAP_global"] >= 70) & (df_actual["CAP_global"] < 85)).sum(), ((df_simulado["CAP_global"] >= 70) & (df_simulado["CAP_global"] < 85)).sum()),
        ("Candidatos alejados", (df_actual["CAP_global"] < 70).sum(), (df_simulado["CAP_global"] < 70).sum()),
    ]
    for etiqueta, actual, simulado in metricas:
        fila += 1
        ws.append([etiqueta, float(actual), float(simulado), float(simulado - actual)])
    _estilizar_tabla(ws, encabezado, fila, 4)
    for row in range(encabezado + 1, fila + 1):
        if ws.cell(row, 1).value == "CAP promedio":
            for col in range(2, 5):
                ws.cell(row, col).number_format = "0.0"
    _ajustar_anchos(ws)

    # Resultados por candidato
    ws = wb.create_sheet("Resultados")
    _escribir_titulo(ws, "Resultados por candidato", "Comparación del escenario actual con la simulación")
    encabezados = [
        "Candidato",
        "CAP archivo (%)",
        "CAP actual (%)",
        "Clasificación actual",
        "CAP simulado (%)",
        "Clasificación simulada",
        "Variación (p.p.)",
    ]
    for comp in competencias:
        encabezados.extend([f"{comp} - actual (%)", f"{comp} - simulado (%)"])
    ws.append([])
    ws.append(encabezados)
    encabezado = 4
    for idx in df_base.index:
        actual = df_actual.at[idx, "CAP_global"]
        simulado = df_simulado.at[idx, "CAP_global"]
        row = [
            df_base.at[idx, "Candidato"],
            df_base.at[idx, "CAP_archivo"],
            actual,
            _clasificacion(actual),
            simulado,
            _clasificacion(simulado),
            simulado - actual,
        ]
        for comp in competencias:
            row.extend([df_actual.at[idx, f"{comp}__cap"], df_simulado.at[idx, f"{comp}__cap"]])
        ws.append(row)
    _estilizar_tabla(ws, encabezado, ws.max_row, len(encabezados))
    for row in range(encabezado + 1, ws.max_row + 1):
        for col in [2, 3, 5, 7] + list(range(8, len(encabezados) + 1)):
            ws.cell(row, col).number_format = "0.0"
    _ajustar_anchos(ws, max_width=32)

    # Configuración final
    ws = wb.create_sheet("Configuración")
    _escribir_titulo(ws, "Configuración de la simulación", "Esperados y pesos utilizados en el cálculo")
    encabezados = [
        "Competencia",
        "Esperado actual",
        "Esperado simulado",
        "Peso final (%)",
        "CAP promedio actual (%)",
        "CAP promedio simulado (%)",
    ]
    ws.append([])
    ws.append(encabezados)
    encabezado = 4
    for comp in competencias:
        ws.append([
            comp,
            esperados_actuales[comp],
            esperados_simulados[comp],
            pesos_simulados[comp],
            df_actual[f"{comp}__cap"].mean(),
            df_simulado[f"{comp}__cap"].mean(),
        ])
    _estilizar_tabla(ws, encabezado, ws.max_row, len(encabezados))
    for row in range(encabezado + 1, ws.max_row + 1):
        for col in [2, 3, 4, 5, 6]:
            ws.cell(row, col).number_format = "0.0"
    _ajustar_anchos(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
