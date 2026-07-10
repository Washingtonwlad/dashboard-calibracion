import unittest
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from calibrador import calcular
from export_excel import crear_excel_resultados


class ExportExcelTests(unittest.TestCase):
    def test_exporta_tres_hojas_con_resultados_y_configuracion(self):
        competencias = ["Liderazgo"]
        base = pd.DataFrame({
            "Candidato": ["Ana"],
            "CAP_archivo": [88.0],
            "Liderazgo__valor": [7.0],
            "Liderazgo__esperado": [8.0],
        })
        actual = calcular(base, competencias, {"Liderazgo": 8}, {"Liderazgo": 1})
        simulado = calcular(base, competencias, {"Liderazgo": 7}, {"Liderazgo": 2})
        data = crear_excel_resultados(
            {"Nombre del Proceso": "Proceso demo"},
            base,
            actual,
            simulado,
            competencias,
            {"Liderazgo": 8},
            {"Liderazgo": 7},
            {"Liderazgo": 2},
        )

        wb = load_workbook(BytesIO(data), data_only=True)
        self.assertEqual(wb.sheetnames, ["Resumen", "Resultados", "Configuración"])
        self.assertEqual(wb["Resultados"]["A5"].value, "Ana")
        self.assertEqual(wb["Configuración"]["C5"].value, 7)
        self.assertEqual(wb["Configuración"]["D5"].value, 2)


if __name__ == "__main__":
    unittest.main()
